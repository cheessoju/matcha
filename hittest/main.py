#-*-coding:utf-8-*-
import openpyxl
import os
import shutil
from openpyxl.styles import PatternFill

yellow=PatternFill(fgColor="FFFF00",fill_type="solid") #黄色背景

#关键词表连接词必须为顿号
fh1='＂＃＄％＆＇（）＊＋，－／；：＜＝＞＠［＼］＾＿｀｛｜｝～｟｠｢｣､\u3000〃〈〉《》「」『』【】〔〕〖〗〘〙〚〛〜〝〞〟〰〾〿–—‘’‛“”„‟…‧﹏﹑﹔·！？｡。'
fh2="!\"#$%&\\'()*,-./:;\u003C=>?@[\\\\]^_`{|}~"
fh=set(fh1+fh2) #用来筛查的标点符号

""" 初始文件夹&保存文件夹独立建立,写绝对路径"""
initial_dir =r"C:\Users\Administrator\Desktop\hittest\input" #初始文件夹路径
save_dir = r"C:\Users\Administrator\Desktop\hittest\output" #保存文件夹路径

def main(initial_dir,save_dir):

    initial_xlsx_list=[i for i in os.listdir(initial_dir)] #获取所有的初始文件路径
    dic={} #对应要保存的路径
    for file in initial_xlsx_list:
        try:
            save_file=file.rstrip(".xlsx")+"_result.xlsx"
            shutil.copy(initial_dir+"/"+file,save_dir+"/"+save_file) #复制文件
            dic[file]=save_file
        except:
            print(f"{file}复制异常,请检查文件是否损坏或者被打开!")

    for k,v in dic.items():
        initial_work=openpyxl.load_workbook(initial_dir+"/"+k)
        save_work=openpyxl.load_workbook(save_dir+"/"+v)


        result_sheet=save_work["关键词表"]
        result_sheet.title="result" #更改名称

        keywordfs_sheet=initial_work["关键词表"]
        order_sheet=initial_work["订单信息表"]

        keywordfs_sheet_rows=keywordfs_sheet.max_row #最大行数
        k_ind=0 #记录“关键字”这几个字所在位置
        keywordfs_list=[] #存储关键字
        for row in range(1,keywordfs_sheet_rows+1):
            v_=keywordfs_sheet[f"C{row}"].value
            if v_=="关键字":
                k_ind=row
                continue
            if k_ind:
                if fh & set(v_):
                    keywordfs_sheet[f"C{row}"].fill=yellow #标黄色
                keywordfs_list.append(v_.split("、"))

        order_list=[] #订单信息列表
        order_sheet_rows=order_sheet.max_row
        order_sheet_cols=order_sheet.max_column

        flag=False
        for row in range(1,order_sheet_rows+1):
            s="" #累计字符串
            for col in range(3,order_sheet_cols+1):
                c=order_sheet.cell(row,col)
                if c.value=="行为分析":
                    flag=True
                if flag:
                    if c.value!=None:
                        s+="       "+c.value 
            if flag:
                order_list.append(s)
        order_list.pop(0)
        print("order_list:",order_list)

        result_list=[] 
        for kw_list in keywordfs_list:
            frequency=0
            for i in order_list:
                for j in kw_list:
                    if j in i: 
                        frequency+=1 
                        
            result_list.append(frequency)

        for i in range(1,k_ind):
            result_sheet.delete_rows(1)
        result_sheet.insert_rows(1)
        result_sheet["C1"].value="关键词验证结果"
        result_sheet["D2"].value="命中数量"
        for i in range(len(result_list)): #写入结果
            result_sheet[f"D{i+3}"].value=result_list[i]

        """保存"""
        initial_work.save(initial_dir + "/" + k)
        save_work.save(save_dir + "/" + v)

main(initial_dir,save_dir)

















